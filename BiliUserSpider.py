from lxml import etree;
import requests;
from openpyxl import Workbook;
import time,re,json

from requests.models import to_key_val_list
from requests.utils import default_user_agent;

reqhead = {
    'Accept':'application/json, text/plain, */*',
    'Accept-Encoding':'gzip, deflate, br',
    'Accept-Language':'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
    'Cache-Control':'no-cache',
    'Connection':'keep-alive',
    'Cookie':"#paste your cookie",
    'Host':'api.bilibili.com',
    'Origin':'https://space.bilibili.com',
    'Pragma':'no-cache',
    'TE':'Trailers',
   'User-Agent':'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:88.0) Gecko/20100101 Firefox/88.0'
}

def getUserInfo(id):
    reqhead['Referer'] = 'https://space.bilibili.com/%d/video'%(id)
    infoPath = 'https://api.bilibili.com/x/space/acc/info?mid=%d&jsonp=jsonp'%(id)
    statPath = 'https://api.bilibili.com/x/relation/stat?vmid=%d&jsonp=jsonp'%(id)
    upstatPath = 'https://api.bilibili.com/x/space/upstat?mid=%d&jsonp=jsonp'%(id)
    info = getJson(infoPath,reqhead)
    stat = getJson(statPath,reqhead)
    upstat = getJson(upstatPath,reqhead)
    return info,stat,upstat

def getVideosPage(id,pn):
    videoPath = 'https://api.bilibili.com/x/space/arc/search?mid=%d&ps=30&tid=0&pn=%d&keyword=&order=pubdate&jsonp=jsonp'%(id,pn)
    page = getJson(videoPath,reqhead)
    return page

def getVideoTags(bvid):
    response = requests.get('https://www.bilibili.com/video/%s'%(bvid))
    response.raise_for_status()
    page = etree.HTML(response.content)
    tags = page.xpath('//ul[@class="tag-area clearfix"]//text()')
    datas = page.xpath('//div[@class="ops"]//text()')
    dlist = []
    tlist = []
    for data in datas:
        result = clean(data)
        if(re.match(r'^\d+$',result,flags=0) != None):
            dlist.append(result)
        else:
            dlist.append('0')
    for tag in tags:
        tlist.append(
            clean(tag)
        )
    return dlist,tlist

def getVedioDetail(bvid):
    response = requests.get('https://www.bilibili.com/video/%s'%(bvid))
    response.raise_for_status()
    page = etree.HTML(response.content)
    scripts = page.xpath('/html/head/script[6]//text()')[0]
    scripts = re.findall(r'(=\{.+\};)',scripts)[0]
    details = json.loads(scripts[1:-1])
    datas = details['videoData']['stat']
    tags = details['tags']
    return datas,tags

def clean(str):
    return re.sub(r'\n|\r|\s','',str,count=0)

def getJson(url,header):
    response = requests.get(url,headers=header)
    response.raise_for_status()
    content = json.loads(response.content.decode(response.encoding))
    if(content['code']==0):
        return content['data']
    else:
        raise Exception('request failed status_code:%d ,message: %s'%(content['code'],content['message']))

__ik= {
    "ID":"mid",
    "昵称":"name",
    "个性签名":"sign",
    "用户等级":"level",
}

__vl = ['bv号','标题','时间戳','发布日期','简介','时长','播放数','弹幕数','评论数','点赞数','投币数','收藏数','转发数','标签','标签id']
__dl = ['view','danmaku','reply','like','coin','favorite','share']

def getUserExcel(id):
    info,stat,upstat= getUserInfo(id)
    wb = Workbook()
    infows = wb.create_sheet(title='简介',index=0)
    index = 1
    for key in __ik: 
        infows.cell(1,index,key)
        infows.cell(2,index,info[__ik[key]])
        index+=1
    infows.cell(1,index,"关注")
    infows.cell(2,index,stat['following'])
    index+=1
    infows.cell(1,index,"粉丝")
    infows.cell(2,index,stat['follower'])
    index+=1
    if(len(upstat)>1):
        infows.cell(1,index,"播放")
        infows.cell(2,index,upstat['archive']['view'])
        index+=1
        infows.cell(1,index,"阅读")
        infows.cell(2,index,upstat['article']['view'])
        index+=1
        infows.cell(1,index,"点赞")
        infows.cell(2,index,upstat['likes'])
    vdfows = wb.create_sheet(title='视频',index=1)
    print('get user info done')
    for i in range(len(__vl)):
        vdfows.cell(1,i+1,__vl[i])
    index = 0
    vpage = getVideosPage(id,1)
    total = vpage['page']['count']
    ps = vpage['page']['ps']
    vlist = vpage['list']['vlist']
    while index < total:
        for vedio in vlist:
            rowindex = index+2
            vdfows.cell(rowindex,1,vedio['bvid'])
            vdfows.cell(rowindex,2,vedio['title'])
            ts = vedio['created']
            date = time.gmtime(ts)
            vdfows.cell(rowindex,3,ts)
            vdfows.cell(rowindex,4,time.strftime("%Y-%m-%d %H:%M:%S", date))
            vdfows.cell(rowindex,5,vedio['description'])
            vdfows.cell(rowindex,6,vedio['length'])
            datas,tags = getVedioDetail(vedio['bvid'])
            for i in range(len(__dl)):
                vdfows.cell(rowindex,7+i,datas[__dl[i]])
            tag = tags.pop()
            tstr = tag['tag_name']
            tids = str(tag['tag_id'])
            for tag in tags:
                tstr = tstr + ' & ' + tag['tag_name']
                tids = tids + ';' + str(tag['tag_id'])
            vdfows.cell(rowindex,14,tstr)
            vdfows.cell(rowindex,15,tids)
            index += 1
            print('get video tags done: %d/%d'%(total,index))
            time.sleep(5)
        if(index<total):
            next = (index/ps)+1
            print('this page done,next page num:%d'%(next))
            vpage = getVideosPage(id,next)
            vlist = vpage['list']['vlist']
    print('get videos done, create file')
    wb.save('%s的视频数据集.xlsx'%(info['name']))